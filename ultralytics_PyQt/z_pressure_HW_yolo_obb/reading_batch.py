from ultralytics import YOLO
import os
import cv2
import pandas as pd
from math import sqrt
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import datetime
import math

# 要检测的图片路径
file_path_all = 'data_show'
# 用于读取Excel文件
file_excel_path = '序号标记对照表.xlsx'
Reading_result_path = 'a_reading_batch_results/'


# 模型路径
model_path1 = 'weight/1biaopan_all/weights/best.pt'
model_path2 = 'weight/2biaopan_nolabel/weights/best.pt'
model_path3 = 'weight/3biaopan_label/weights/best.pt'
model_path4 = 'weight/4read/weights/best.pt'
# 输出路径    改result_path
current_time = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
current_time2 = datetime.datetime.now().strftime("%Y-%m-%d")
result_path = 'outputs/outputs-' + current_time + '/'
result_path1 = result_path + 'all/'
result_path2_1 = result_path + 'biaopan/'
result_path2_2 = result_path + 'biaoqian/'
txt_path1 = result_path + 'biaoqian.txt'
txt_path2 = result_path + 'dushu.txt'
pic_path = result_path + 'fitcenter/'


# 读取Excel文件
df = pd.read_excel(file_excel_path, engine='openpyxl')
# 将序号列的数据类型转换为字符串
df['序号'] = df['序号'].astype(str)
# Excel文件中第一列是序号，第二列是表计 提取这两列并转换为字典
my_dict = pd.Series(df['表计'].values,index=df['序号']).to_dict()

def makeFiledir(result_path):
    """ 创建输出文件夹"""
    if not os.path.exists(result_path):  # 是否存在这个文件夹
        os.makedirs(result_path)  # 如果没有这个文件夹，那就创建一个


def calculate_intersection(p1, p2, p3, p4):
    # 计算第一条直线的斜率和截距
    m1 = (p2[1] - p1[1]) / (p2[0] - p1[0])
    b1 = p1[1] - m1 * p1[0]

    # 计算第二条直线的斜率和截距
    m2 = (p4[1] - p3[1]) / (p4[0] - p3[0])
    b2 = p3[1] - m2 * p3[0]

    # 检查斜率是否相同，即直线是否平行
    if m1 == m2:
        return None  # 直线平行或重合，没有交点

    # 计算交点的x坐标
    x = (b2 - b1) / (m1 - m2)
    # 计算交点的y坐标
    y = m1 * x + b1

    return (x, y)

def Distances(a, b):
    # 返回两点间的距离
    x1, y1 = a
    x2, y2 = b
    Distances = int(sqrt((x1 - x2) ** 2 + (y1 - y2) ** 2))
    return Distances

def GetClockAngle(v1, v2):
    # 2个向量模的乘积 ,返回夹角
    TheNorm = np.linalg.norm(v1) * np.linalg.norm(v2)
    # 叉乘
    rho = np.rad2deg(np.arcsin(np.cross(v1, v2) / TheNorm))
    # 点乘
    theta = np.rad2deg(np.arccos(np.dot(v1, v2) / TheNorm))
    if rho > 0:
        return theta
    else:
        return 360 - theta

def get_mid_point(x_p1, y_p1, x_p2, y_p2, x_p3, y_p3, x_p4, y_p4):
    # 定义点
    P1 = (x_p1, y_p1)
    P2 = (x_p2, y_p2)
    P3 = (x_p3, y_p3)
    P4 = (x_p4, y_p4)
    # 边及其对应的两个点
    edges = [
        (P1, P2),
        (P2, P3),
        (P3, P4),
        (P4, P1)
    ]

    # 计算边长并排序
    edges_with_length = [((p1, p2), math.hypot(p2[0] - p1[0], p2[1] - p1[1])) for (p1, p2) in edges]
    edges_with_length.sort(key=lambda x: x[1])  # 按边长升序排列

    # 获取两个最短边的中点
    def midpoint(p1, p2):
        return ((p1[0] + p2[0]) / 2, (p1[1] + p2[1]) / 2)

    short_edge_1 = edges_with_length[0][0]
    short_edge_2 = edges_with_length[1][0]

    midpoint1 = midpoint(*short_edge_1)
    midpoint2 = midpoint(*short_edge_2)

    return midpoint1, midpoint2

# 创建路径
makeFiledir(result_path)
makeFiledir(result_path1)
makeFiledir(pic_path)
makeFiledir(Reading_result_path)

# 遍历文件夹
for filename in os.listdir(file_path_all):
    # 第一步将表盘带标签裁剪出来
    yolo1 = YOLO(model=model_path1, task='detect')  # task非必须
    result1 = yolo1(source=file_path_all + '/' + filename, save=True, save_txt=True, save_crop=True, conf=0.7)
    save_path1 = result1[0].save_dir
    # txt文件路径
    label_path1 = save_path1 + '/' + 'labels'
    # 保存表盘带标签
    image1 = cv2.imread(save_path1 + '/crops/Instrument/' + filename)
    cv2.imwrite(result_path1 + filename, image1)

    # 检查文件是否是图片，这里假设图片扩展名为.jpg, .jpeg, .png, .bmp之一
    if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')):
        # 拼接完整的文件路径
        imagepath = os.path.join(result_path1, filename)
        image_name = imagepath.split('/')[-1]
        # 第二步从第一步得到的图片中将表盘裁剪出来
        yolo2 = YOLO(model=model_path2, task='detect')  # task非必须
        result2 = yolo2(source=imagepath, save=True, save_txt=True, save_crop=True, conf=0.7)     # imagepath表盘带标签图像
        save_path2 = result2[0].save_dir
        # txt文件路径
        label_path2 = save_path2 + '/' + 'labels'
        print("label_path2:", label_path2)
        result_path2 = result_path2_1 + image_name.split('.')[0] + '/'
        makeFiledir(result_path2)
        # 保存单表盘
        image2 = cv2.imread(save_path2 + '/crops/Pointer/' + filename)
        cv2.imwrite(result_path2 + filename, image2)

        # 第三步从第一步得到的图片中将标签裁剪出来
        yolo3 = YOLO(model=model_path3, task='detect')  # task非必须
        result3 = yolo3(source=imagepath, save=True, save_txt=True, save_crop=True, conf=0.7)    # imagepath表盘带标签图像
        save_path3 = result3[0].save_dir
        # txt文件路径
        label_path3 = save_path3 + '/' + 'labels'
        print("label_path3:", label_path3)
        result_path3 = result_path2_2 + image_name.split('.')[0] + '/'
        makeFiledir(result_path3)
        # 保存标签
        image3 = cv2.imread(save_path3 + '/crops/Label/' + filename)
        cv2.imwrite(result_path3 + filename, image3)


        # 第四步识别标签
        yolo4 = YOLO(model=model_path4, task='detect')  # task非必须
        result4 = yolo4(source=result_path3, save=True, save_txt=True, conf=0.7)
        save_path4 = result4[0].save_dir
        # txt文件路径
        label_path4 = save_path4 + '/' + 'labels'
        print("label_path4:", label_path4)

        # 初始化一个空列表，用于存储文件的第一列数据
        all_data = []

        # 遍历文件夹中的所有txt文件
        for filename in os.listdir(label_path4):
            if filename.endswith('.txt'):
                file_path = os.path.join(label_path4, filename)
                with open(file_path, 'r') as file:
                    for line in file:
                        first_column = line.split()[0]  # 读取每行的第一列
                        with open(txt_path1, 'a', encoding='UTF-8') as file:
                            file.write(image_name.split('.')[0])
                            file.write(' ')
                            file.write(my_dict[first_column] + '\n')


        # 第五步识别读数
        yolo5 = YOLO(model=model_path4, task='detect')  # task非必须
        result5 = yolo5(source=result_path2, save=True, save_txt=True, conf=0.7)    # result_path2表盘图像
        save_path5 = result5[0].save_dir
        # txt文件路径
        label_path5 = save_path5 + '/' + 'labels'
        print("label_path5:", label_path5)

        # 读取图像
        image = cv2.imread(save_path5 + '/' + image_name)
        print(save_path5 + '/' + image_name)
        # 获取图像的高度和宽度
        height, width = image.shape[:2]
        print(height, width)
        # height, width = result5[0].orig_shape

        # 遍历文件夹中的所有txt文件
        for filename in os.listdir(label_path5):
            if filename.endswith('.txt'):
                file_path = os.path.join(label_path5, filename)
                # 初始化一个空列表，用于存储每一行的数据
                rows_as_lists = []
                with open(file_path, 'r') as f:
                    for line in f:
                        # 使用split()方法将行分割成列表，这里假设数据之间是用空格分隔的
                        row_list = line.strip().split()  # strip()用于移除字符串头尾指定的字符（默认为空格）
                        # 将得到的列表添加到rows_as_lists中
                        rows_as_lists.append(row_list)

                # 先按第一列排序，再按第二列排序
                sorted_rows = sorted(rows_as_lists, key=lambda x: (float(x[0]), float(x[1])))
                # print(sorted_rows)

                # 起始刻度线坐标
                x_s1 = width * float(sorted_rows[0][7])
                y_s1 = height * float(sorted_rows[0][8])
                x_s2 = width * float(sorted_rows[0][1])
                y_s2 = height * float(sorted_rows[0][2])
                x_s3 = width * float(sorted_rows[0][3])
                y_s3 = height * float(sorted_rows[0][4])
                x_s4 = width * float(sorted_rows[0][5])
                y_s4 = height * float(sorted_rows[0][6])
                # 拟合起始刻度线坐标
                x_sf = (x_s1 + x_s2 + x_s3 + x_s4) / 4
                y_sf = (y_s1 + y_s2 + y_s3 + y_s4) / 4

                # 终止刻度线坐标
                x_e1 = width * float(sorted_rows[1][1])
                y_e1 = height * float(sorted_rows[1][2])
                x_e2 = width * float(sorted_rows[1][3])
                y_e2 = height * float(sorted_rows[1][4])
                x_e3 = width * float(sorted_rows[1][5])
                y_e3 = height * float(sorted_rows[1][6])
                x_e4 = width * float(sorted_rows[1][7])
                y_e4 = height * float(sorted_rows[1][8])
                # 拟合终止刻度线坐标
                x_ef = (x_e1 + x_e2 + x_e3 + x_e4) / 4
                y_ef = (y_e1 + y_e2 + y_e3 + y_e4) / 4

                # 指针端点坐标
                x_p1 = width * float(sorted_rows[3][7])
                y_p1 = height * float(sorted_rows[3][8])
                x_p2 = width * float(sorted_rows[3][1])
                y_p2 = height * float(sorted_rows[3][2])
                x_p3 = width * float(sorted_rows[3][3])
                y_p3 = height * float(sorted_rows[3][4])
                x_p4 = width * float(sorted_rows[3][5])
                y_p4 = height * float(sorted_rows[3][6])
                # 拟合指针端点坐标
                x_pf = (x_p1 + x_p2 + x_p3 + x_p4) / 4
                y_pf = (y_p1 + y_p2 + y_p3 + y_p4) / 4

                # 中间刻度线坐标
                x_z1 = width * float(sorted_rows[2][1])
                y_z1 = height * float(sorted_rows[2][2])
                x_z2 = width * float(sorted_rows[2][3])
                y_z2 = height * float(sorted_rows[2][4])
                x_z3 = width * float(sorted_rows[2][5])
                y_z3 = height * float(sorted_rows[2][6])
                x_z4 = width * float(sorted_rows[2][7])
                y_z4 = height * float(sorted_rows[2][8])
                # 拟合中间刻度线坐标
                x_zf = (x_z1 + x_z2 + x_z3 + x_z4) / 4
                y_zf = (y_z1 + y_z2 + y_z3 + y_z4) / 4

                # 计算圆心坐标
                # 两个短边中点坐标
                (x_ss1, y_ss1), (x_ss2, y_ss2) = get_mid_point(x_s1, y_s1, x_s2, y_s2, x_s3, y_s3, x_s4, y_s4)
                p1 = (x_ss1, y_ss1)
                p2 = (x_ss2, y_ss2)

                # 两个短边中点坐标
                (x_ee1, y_ee1), (x_ee2, y_ee2) = get_mid_point(x_e1, y_e1, x_e2, y_e2, x_e3, y_e3, x_e4, y_e4)
                p3 = (x_ee1, y_ee1)
                p4 = (x_ee2, y_ee2)

                # 两个短边中点坐标
                (x_pp1, y_pp1), (x_pp2, y_pp2) = get_mid_point(x_p1, y_p1, x_p2, y_p2, x_p3, y_p3, x_p4, y_p4)
                p5 = (x_pp1, y_pp1)
                p6 = (x_pp2, y_pp2)

                # 两个短边中点坐标
                (x_zz1, y_zz1), (x_zz2, y_zz2) = get_mid_point(x_z1, y_z1, x_z2, y_z2, x_z3, y_z3, x_z4, y_z4)
                p7 = (x_zz1, y_zz1)
                p8 = (x_zz2, y_zz2)


                # 计算开始刻度线、结束刻度线、中间刻度线交点坐标
                intersection_se = calculate_intersection(p1, p2, p3, p4)
                intersection_sz = calculate_intersection(p1, p2, p7, p8)
                intersection_ez = calculate_intersection(p3, p4, p7, p8)

                d11 = Distances(intersection_se, intersection_sz)
                d22 = Distances(intersection_se, intersection_ez)
                d33 = Distances(intersection_sz, intersection_ez)
                d44 = min(d11, d22, d33)

                # 由开始刻度线、结束刻度线、中间刻度线计算的圆心坐标1
                if d44 == d11:
                    c_xx = (intersection_se[0] + intersection_sz[0]) / 2
                    c_yy = (intersection_se[1] + intersection_sz[1]) / 2
                elif d44 == d22:
                    c_xx = (intersection_se[0] + intersection_ez[0]) / 2
                    c_yy = (intersection_se[1] + intersection_ez[1]) / 2
                elif d44 == d33:
                    c_xx = (intersection_sz[0] + intersection_ez[0]) / 2
                    c_yy = (intersection_sz[1] + intersection_ez[1]) / 2

                # 计算计算开始刻度线、结束刻度线、指针交点坐标
                intersection_se = calculate_intersection(p1, p2, p3, p4)
                intersection_sp = calculate_intersection(p1, p2, p5, p6)
                intersection_ep = calculate_intersection(p3, p4, p5, p6)

                d1 = Distances(intersection_se, intersection_sp)
                d2 = Distances(intersection_se, intersection_ep)
                d3 = Distances(intersection_sp, intersection_ep)
                d4 = min(d1, d2, d3)

                # 由开始刻度线、结束刻度线、指针计算的圆心坐标2
                if d4 == d1:
                    c_x = (intersection_se[0] + intersection_sp[0]) / 2
                    c_y = (intersection_se[1] + intersection_sp[1]) / 2
                elif d4 == d2:
                    c_x = (intersection_se[0] + intersection_ep[0]) / 2
                    c_y = (intersection_se[1] + intersection_ep[1]) / 2
                elif d4 == d3:
                    c_x = (intersection_sp[0] + intersection_ep[0]) / 2
                    c_y = (intersection_sp[1] + intersection_ep[1]) / 2

                # 定义两个点的坐标
                point1 = (int(intersection_se[0]), int(intersection_se[1]))
                point2 = (int(intersection_sp[0]), int(intersection_sp[1]))
                point3 = (int(intersection_ep[0]), int(intersection_ep[1]))
                point4 = (int(c_x), int(c_y))
                point5 = (int(x_zf), int(y_zf))
                point6 = (int(c_xx), int(c_yy))

                # 在图像上标记点，这里设置圆的半径为5，颜色为红色(0, 0, 255)，线宽为-1表示填充圆
                cv2.circle(image, point1, 5, (0, 0, 255), -1)  # 开始 结束
                cv2.circle(image, point2, 5, (0, 0, 120), -1)  # 开始 指针
                cv2.circle(image, point3, 5, (0, 0, 0), -1)  # 结束 指针
                cv2.circle(image, point4, 5, (0, 0, 255), -1)
                cv2.circle(image, point5, 5, (0, 0, 255), -1)
                cv2.circle(image, point6, 5, (0, 0, 255), -1)

                # 保存图像
                cv2.imwrite(pic_path + image_name + "_fitting.jpg", image)


                # 计算修正前读数
                # 计算零刻度与指针端点所成角度
                v1 = [x_sf - c_x, y_sf - c_y]
                v2 = [x_pf - c_x, y_pf - c_y]
                theta1 = GetClockAngle(v1, v2)
                # 计算零刻度与最终刻度所成角度
                v3 = [x_sf - c_x, y_sf - c_y]
                v4 = [x_ef - c_x, y_ef - c_y]
                theta2 = GetClockAngle(v3, v4)
                # 计算修正前最终读数
                readValue1 = theta1 / theta2 * 1 - 0.1
                # print(readValue1)

                # 计算修正后读数
                # 最终圆心坐标
                cx = (c_x + c_xx) / 2
                cy = (c_y + c_yy) / 2
                # 计算零刻度与指针端点所成角度
                v5 = [x_sf - cx, y_sf - cy]
                v6 = [x_pf - cx, y_pf - cy]
                theta3 = GetClockAngle(v5, v6)

                # 计算零刻度与最终刻度所成角度
                v7 = [x_sf - cx, y_sf - cy]
                v8 = [x_ef - cx, y_ef - cy]
                theta4 = GetClockAngle(v7, v8)

                # 计算最终读数
                readValue2 = theta3 / theta4 * 1 - 0.1
                # print(readValue2)

                # 计算零刻度与0.4所成角度
                v9 = [x_sf - cx, y_sf - cy]
                v10 = [x_zf - cx, y_zf - cy]
                theta5 = GetClockAngle(v9, v10)
                # 计算最终读数2
                readValue3 = theta5 / theta4 * 1 - 0.1
                # print(readValue3)

                current_time = datetime.datetime.now().strftime("%Y-%m-%d")
                with open(txt_path2, 'a') as file:
                    file.write(image_name.split('.')[0] + ' ')
                    file.write(str(readValue2 + (0.4 - readValue3) / 2) + ' ')
                    file.write(current_time + '\n')



# 读取文本文件的函数
def read_txt_file(file_name):
    encodings = ['utf-8', 'gbk', 'latin1', 'ISO-8859-1', ]  # 常见的一些编码
    for encoding in encodings:
        try:
            with open(file_name, 'r', encoding=encoding) as file:
                column1, column2 = [], []
                for line in file:
                    parts = line.strip().split()  # 假设数据是空格分隔的
                    if len(parts) >= 2:
                        column1.append(parts[0])
                        column2.append(parts[1])
                return column1, column2
        except UnicodeDecodeError:
            continue  # 如果当前编码解码失败，尝试下一个编码
    raise ValueError(f'所有尝试的编码都无法读取{file_name}')


# 读取biaoqian.txt和dushu.txt文件
biaoqian_col1, biaoqian_col2 = read_txt_file(txt_path1)
# dushu_col1, dushu_col2 = read_txt_file('outputs2/dushu.txt')

# 创建一个新的工作簿
wb = Workbook()
ws = wb.active

# 写入标题行 图片名称 表计 最终读数
ws.append(["图片名称", "表计", "最终读数", "检测时间"])

# 写入biaoqian.txt的数据
for i in range(len(biaoqian_col1)):
    print(biaoqian_col1[i], biaoqian_col2[i])
    ws.append([biaoqian_col1[i], biaoqian_col2[i]])

# # 写入dushu.txt的第二列数据到第三列下面，从第二行开始
# for i, value in enumerate(dushu_col2, start=2):  # 从第二行开始
#     ws.cell(row=i, column=3, value=value)

# 保存工作簿为一个Excel文件
wb.save(result_path + "Reading results-" + current_time2 + '.xlsx')


# 读取Excel文件
excel_path = result_path + "Reading results-" + current_time2 + '.xlsx'

df_excel = pd.read_excel(excel_path)

# 读取txt文件
df_txt = pd.read_csv(txt_path2, sep=' ', header=None, names=['图片名称', '读数', '检测时间'])

# 将txt文件的数据映射到Excel表中
df_excel['最终读数'] = df_excel['图片名称'].map(df_txt.set_index('图片名称')['读数'])

# 对于没有匹配到的，填充"读数失败"
df_excel['最终读数'].fillna('读数失败', inplace=True)

# 将txt文件的检测时间数据映射到Excel表中
df_excel['检测时间'] = df_excel['图片名称'].map(df_txt.set_index('图片名称')['检测时间'])

# 对于没有匹配到的检测时间留空
df_excel['检测时间'].fillna('未知', inplace=True)

# 保存更新后的Excel文件
df_excel.to_excel(excel_path, index=False)

# 使用openpyxl加载这个Excel文件
wb = load_workbook(excel_path)
ws = wb.active  # 激活工作表

# 遍历所有单元格，设置它们的对齐方式为居中
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 列宽设置
column_widths = [15, 30, 20, 20]  # 示例中只有四列，根据实际列数调整

# 遍历column_widths列表来设置每一列的宽度
for i, column_width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = column_width
# 保存修改后的Excel文件
wb.save(excel_path)

# 加载已有的Excel文件
wb = load_workbook(excel_path)

new_file_path = Reading_result_path + "Reading results-" + current_time2 + '.xlsx'

# 直接保存到新的路径
wb.save(new_file_path)

print("数据已成功写入Excel文件。")
